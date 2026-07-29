"""
Consulta el tiempo de entrega estimado de un producto en Amazon.es
para varios códigos postales y lo vuelca a un Excel con histórico.

Pensado para ejecutarse dos veces al día (tarea programada).

Requisitos (una sola vez):
    pip install playwright openpyxl
    playwright install chromium

Uso:
    python main.py
"""

import os
import sys
import time
import random
import subprocess
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from playwright.sync_api import sync_playwright, TimeoutError as PWTimeout
from playwright.sync_api import expect

# ─────────────────────────────────────────────────────────────
# CONFIGURACIÓN — esto es lo único que tocas normalmente
# ─────────────────────────────────────────────────────────────

# URL del producto en Amazon.es (pídesela a tu compañera)
if len(sys.argv) != 2:
    print("ERROR: No se ha indicado ninguna URL.")
    print("Ejecuta el programa desde ejecutar.bat")
    sys.exit(1)

PRODUCT_URL = sys.argv[1]

# Un código postal representativo por provincia (19).
# "provincia": "código postal". Cambia los que quiera ella.
CODIGOS_POSTALES = {
    "A Coruña": "15001",
    "Madrid": "28001",
    "Barcelona": "08001",
    "Valencia": "46001",
    "Sevilla": "41001",
    "Zaragoza": "50001",
    "Málaga": "29001",
    "Murcia": "30001",
    "Bilbao (Vizcaya)": "48001",
    "Alicante": "03001",
    "Valladolid": "47001",
    "Vigo (Pontevedra)": "36201",
    "Granada": "18001",
    "Oviedo (Asturias)": "33001",
    "Santa Cruz de Tenerife": "38001",
    "Las Palmas": "35001",
    "Palma (Baleares)": "07001",
    "Córdoba": "14001",
    "Valencia (interior)": "46021",
}

# Archivo Excel de salida (se crea si no existe, se amplía si existe)
EXCEL_PATH = Path("tiempos_envio.xlsx")

# Ejecutar sin ventana visible. Ponlo en False la primera vez para VER
# lo que hace y comprobar que funciona.
HEADLESS = False

# ─────────────────────────────────────────────────────────────
# LÓGICA
# ─────────────────────────────────────────────────────────────

def consultar_codigo_postal(page, codigo_postal: str) -> str:
    try:

        page.keyboard.press("Escape")
        time.sleep(0.5)
        # Abre el modal de ubicación
        page.click("#nav-global-location-popover-link", timeout=15000)

        # Espera a que el modal esté abierto
        page.locator("#GLUXZipUpdateInput:visible, #GLUXChangePostalCodeLink:visible").first.wait_for()

        change = page.locator("#GLUXChangePostalCodeLink:visible")
        zip_input = page.locator("#GLUXZipUpdateInput:visible")

        # Dejamos terminar la animación del modal
        page.wait_for_timeout(800)

        # Primera ejecución: el campo ya está visible.
        # Siguientes ejecuciones: hay que pulsar "Cambiar".
        if zip_input.count() == 0:
            if change.count() == 0:
                raise Exception("No aparece ni el campo del código postal ni el botón 'Cambiar'.")

            change.click(timeout=10000)
            zip_input.wait_for(state="visible", timeout=10000)

        # Rellenar el código postal
        zip_input.fill(codigo_postal)

        # Confirmar el CP (el botón visible)
        page.locator("#GLUXZipUpdate:visible input[type='submit']").click()

        # Esperar a que el CP confirmado sea el que introdujimos (también el visible)
        expect(page.locator("#GLUXZipConfirmationValue")).to_contain_text(
            codigo_postal, timeout=10000
        )

        # Cerrar el modal con "Hecho"
        page.locator("button[name='glowDoneButton']:visible").click()

        # Dar tiempo a que la página actualice la estimación
        time.sleep(random.uniform(2.5, 4.0))


        # ── SELECTOR FRÁGIL: la fecha de entrega ──
        for sel in [
            "#mir-layout-DELIVERY_BLOCK-slot-PRIMARY_DELIVERY_MESSAGE_LARGE .a-text-bold",
            "#deliveryBlockMessage .a-text-bold",
            "#delivery-block-DELIVERY_offer .a-text-bold",
            "[data-csa-c-delivery-time]",
        ]:
            el = page.query_selector(sel)
            if el and el.inner_text().strip():
                return el.inner_text().strip()

        return "ERROR: no se encontró la fecha de entrega (revisar selectores)"

    except PWTimeout as e:
        return f"ERROR: timeout ({e.__class__.__name__})"
    except Exception as e:
        return f"ERROR: {e}"


def guardar_en_excel(filas: list[dict]):
    """Añade filas al Excel (histórico). Crea el archivo si no existe."""
    cabeceras = ["fecha_hora", "provincia", "codigo_postal", "estimacion_entrega"]

    if EXCEL_PATH.exists():
        wb = load_workbook(EXCEL_PATH)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "tiempos"
        ws.append(cabeceras)

    for fila in filas:
        ws.append([fila[c] for c in cabeceras])

    wb.save(EXCEL_PATH)


def abrir_excel():
    """Abre el Excel con el programa por defecto, para que ella no tenga que buscarlo."""
    try:
        ruta = str(EXCEL_PATH.resolve())
        if sys.platform.startswith("win"):
            os.startfile(ruta)  # Windows
        elif sys.platform == "darwin":
            subprocess.run(["open", ruta])  # Mac
        else:
            subprocess.run(["xdg-open", ruta])  # Linux
    except Exception as e:
        print(f"(No se pudo abrir el Excel automáticamente: {e})")


def main():
    momento = datetime.now().strftime("%Y-%m-%d %H:%M")
    print(f"[{momento}] Iniciando consulta de {len(CODIGOS_POSTALES)} códigos postales…")

    filas = []
    errores = 0

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=HEADLESS, slow_mo=300)
        # Un contexto con user-agent normal, para parecer un navegador cualquiera
        context = browser.new_context(
            locale="es-ES",
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/125.0.0.0 Safari/537.36"
            ),
        )
        page = context.new_page()
        page.goto(PRODUCT_URL, wait_until="domcontentloaded", timeout=30000)
        time.sleep(random.uniform(2.0, 3.5))

        try:
            page.click("#sp-cc-rejectall-link", timeout=5000)
            time.sleep(1)
        except PWTimeout:
            pass  # si no sale el banner, seguimos

        for provincia, cp in CODIGOS_POSTALES.items():
            estimacion = consultar_codigo_postal(page, cp)
            if estimacion.startswith("ERROR"):
                errores += 1
            print(f"  {provincia:28} ({cp}): {estimacion}")
            filas.append({
                "fecha_hora": momento,
                "provincia": provincia,
                "codigo_postal": cp,
                "estimacion_entrega": estimacion,
            })
            # Pausa entre consultas: parecer humano y no machacar Amazon
            time.sleep(random.uniform(3.0, 6.0))

        #browser.close()

    guardar_en_excel(filas)
    print(f"Guardado en {EXCEL_PATH.resolve()}")
    abrir_excel()

    # Si TODO falló, salir con error para que la tarea programada avise
    if errores == len(CODIGOS_POSTALES):
        print("¡ATENCIÓN! Han fallado TODAS las consultas. "
              "Probablemente Amazon cambió el HTML: revisa los selectores.")
        sys.exit(1)
    elif errores:
        print(f"Aviso: {errores} de {len(CODIGOS_POSTALES)} consultas fallaron.")


if __name__ == "__main__":
    main()